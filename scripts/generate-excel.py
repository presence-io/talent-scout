#!/usr/bin/env python3
"""Generate HR-ready Excel from talent-scout shortlist + evaluation data."""

import json
import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip3 install openpyxl")
    sys.exit(1)


def load_data(eval_dir: str):
    with open(f"{eval_dir}/shortlist.json") as f:
        shortlist = json.load(f)
    with open(f"{eval_dir}/evaluation.json") as f:
        evaluation = json.load(f)
    return shortlist, evaluation


def build_candidate_context(entry, eval_data):
    """Build a rich context string for AI to generate summary + email."""
    username = entry["username"]
    full = eval_data.get(username, {})
    profile = full.get("profile", {})
    repos = profile.get("recent_repos", [])[:5]
    evaluation = full.get("evaluation", {})

    top_repos = []
    for r in repos:
        if not r.get("is_fork") and r.get("stars", 0) > 0:
            top_repos.append(f"{r['name']} (★{r['stars']}, {r.get('language', '?')}): {(r.get('description') or 'no desc')[:80]}")

    context = f"""Username: {username}
Name: {entry.get('name', '?')}
City: {entry.get('city', '?')}
Company: {entry.get('company', '?')}
Bio: {(profile.get('bio') or '')[:200]}
Followers: {profile.get('followers', 0)}
Public repos: {profile.get('public_repos', 0)}
Score: {entry.get('final_score', 0):.1f}
AI Tier: {entry.get('ai_depth_tier', '?')}
Skill evidence: {', '.join(evaluation.get('skill_evidence', []))}
AI evidence: {', '.join(evaluation.get('ai_depth_evidence', []))}
Top repos:
{chr(10).join(top_repos) if top_repos else 'None'}
Recommended action: {entry.get('recommended_action', '?')}"""
    return context


def generate_ai_content(candidates_context: list[tuple[str, str]]) -> dict:
    """Call claude -p to generate summaries + emails for a batch."""
    batch_input = []
    for username, ctx in candidates_context:
        batch_input.append({"username": username, "context": ctx})

    prompt = f"""You are an HR recruiter writing outreach materials for top developer candidates.

For each candidate below, generate:
1. "why" — 1-2 sentences in Chinese explaining why we found this person and why they're interesting (focus on concrete projects, star counts, AI tool usage)
2. "email_subject" — An attractive email subject line in Chinese that would make a developer want to open it (personalized, mention their project or skill)
3. "email_body" — A short email body in Chinese (3-4 sentences) for cold outreach. Be warm, specific about their work, and mention we're looking for AI-native developers. End with a call to action.

Candidates:
{json.dumps(batch_input, ensure_ascii=False, indent=2)}

Return ONLY valid JSON:
{{"results": [{{"username": "...", "why": "...", "email_subject": "...", "email_body": "..."}}]}}"""

    result = subprocess.run(
        ["claude", "-p", prompt, "--output-format", "text", "--model", "sonnet"],
        capture_output=True, text=True, timeout=180
    )
    stdout = result.stdout.strip()
    # Strip markdown fences if present
    if stdout.startswith("```"):
        stdout = stdout.split("\n", 1)[1] if "\n" in stdout else stdout
        if stdout.endswith("```"):
            stdout = stdout[:-3].strip()
    return json.loads(stdout)


def main():
    eval_dir = sys.argv[1] if len(sys.argv) > 1 else "workspace-data/output/evaluated/2026-03-29T0704"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "workspace-data/talent-scout-shortlist.xlsx"

    print(f"Loading data from {eval_dir}...")
    shortlist, evaluation = load_data(eval_dir)

    # Sort by final_score descending
    shortlist.sort(key=lambda x: x.get("final_score", 0), reverse=True)

    # Only process candidates with recommended_action = reach_out or monitor with score >= 6.0
    top_candidates = [
        c for c in shortlist
        if c.get("recommended_action") == "reach_out" or c.get("final_score", 0) >= 6.0
    ]
    rest = [c for c in shortlist if c not in top_candidates]

    print(f"Total shortlist: {len(shortlist)}")
    print(f"Top candidates for AI content: {len(top_candidates)}")

    # Generate AI content in batches of 10
    ai_content = {}
    batch_size = 10
    for i in range(0, len(top_candidates), batch_size):
        batch = top_candidates[i:i + batch_size]
        batch_ctx = [(c["username"], build_candidate_context(c, evaluation)) for c in batch]
        batch_num = i // batch_size + 1
        total_batches = (len(top_candidates) + batch_size - 1) // batch_size
        print(f"  Generating content batch {batch_num}/{total_batches} ({len(batch)} candidates)...")

        try:
            result = generate_ai_content(batch_ctx)
            for r in result.get("results", []):
                ai_content[r["username"]] = r
        except Exception as e:
            print(f"  Warning: batch {batch_num} failed: {e}")
            for c in batch:
                ai_content[c["username"]] = {
                    "why": "",
                    "email_subject": "",
                    "email_body": ""
                }

    # Create Excel
    print("Creating Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Talent Shortlist"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    reach_out_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    monitor_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        "推荐级别", "GitHub ID", "姓名", "城市", "公司",
        "综合评分", "技术评分", "AI深度评分", "AI深度级别",
        "为什么关注TA", "邮件标题", "邮件内容",
        "Email", "博客", "GitHub主页",
        "信号类型", "信号数量"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    row = 2
    all_candidates = top_candidates + rest
    for c in all_candidates:
        username = c["username"]
        ai = ai_content.get(username, {})
        action = c.get("recommended_action", "skip")
        action_label = {"reach_out": "★ 推荐联系", "monitor": "关注", "skip": "跳过"}.get(action, action)

        values = [
            action_label,
            username,
            c.get("name", ""),
            c.get("city", ""),
            c.get("company", ""),
            round(c.get("final_score", 0), 1),
            round(c.get("skill_score", 0), 1),
            round(c.get("ai_depth_score", 0), 1),
            {"builder": "AI Builder", "user": "AI User", "consumer": "AI Consumer", "none": "无"}.get(
                c.get("ai_depth_tier", ""), c.get("ai_depth_tier", "")
            ),
            ai.get("why", ""),
            ai.get("email_subject", ""),
            ai.get("email_body", ""),
            c.get("email", ""),
            c.get("blog", ""),
            c.get("profile_url", ""),
            ", ".join(c.get("signal_types", [])),
            c.get("signal_count", 0),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if action == "reach_out":
                cell.fill = reach_out_fill
            elif action == "monitor" and c.get("final_score", 0) >= 6.5:
                cell.fill = monitor_fill

        row += 1

    # Column widths
    col_widths = {
        1: 14, 2: 18, 3: 16, 4: 10, 5: 14,
        6: 10, 7: 10, 8: 10, 9: 12,
        10: 40, 11: 30, 12: 50,
        13: 25, 14: 25, 15: 30,
        16: 18, 17: 10,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:Q{row - 1}"

    wb.save(output_path)
    print(f"✓ Excel saved: {output_path}")
    print(f"  {len(all_candidates)} candidates, {len(ai_content)} with AI-generated content")
    return output_path


if __name__ == "__main__":
    main()
